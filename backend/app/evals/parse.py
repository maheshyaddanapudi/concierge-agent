"""Eval upload parsing (spec §15): csv/xlsx in the predefined format —
columns `level (skill|sub_agent)`, `target_id`, `input`, `expected`,
`judge_notes` (optional), `grader` (optional, default llm_judge). One file
targets ONE definition; mixed targets are rejected."""

import csv
import io
from typing import Any
from uuid import UUID

VALID_LEVELS = {"skill", "sub_agent"}
VALID_GRADERS = {"exact", "contains", "llm_judge"}
REQUIRED_COLUMNS = {"level", "target_id", "input", "expected"}


class EvalParseError(ValueError):
    """The upload does not match the predefined format."""


def _rows_from_csv(data: bytes) -> list[dict[str, str]]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [{(k or "").strip().lower(): (v or "").strip() for k, v in row.items()} for row in reader]


def _rows_from_xlsx(data: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook  # type: ignore[attr-defined]

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)  # type: ignore[no-untyped-call]
    ws = wb.active
    if ws is None:
        return []
    it = ws.iter_rows(values_only=True)
    try:
        header = [str(c or "").strip().lower() for c in next(it)]
    except StopIteration:
        return []
    rows = []
    for raw in it:
        if raw is None or all(c is None for c in raw):
            continue
        rows.append(
            {
                header[i]: str(raw[i] if i < len(raw) and raw[i] is not None else "").strip()
                for i in range(len(header))
            }
        )
    return rows


def parse_eval_file(filename: str, data: bytes) -> dict[str, Any]:
    """Returns {level, target_id, cases: [{input, expected, judge_notes,
    grader}]}. Raises EvalParseError with a actionable message."""
    name = filename.lower()
    if name.endswith(".xlsx"):
        rows = _rows_from_xlsx(data)
    elif name.endswith(".csv"):
        rows = _rows_from_csv(data)
    else:
        raise EvalParseError("upload a .csv or .xlsx file")
    if not rows:
        raise EvalParseError("the file has no data rows")
    missing = REQUIRED_COLUMNS - set(rows[0].keys())
    if missing:
        raise EvalParseError(
            f"missing column(s) {sorted(missing)} — expected "
            "level,target_id,input,expected[,judge_notes][,grader]"
        )
    level: str | None = None
    target_id: str | None = None
    cases: list[dict[str, str]] = []
    for i, row in enumerate(rows, start=2):
        row_level = row.get("level", "")
        if row_level not in VALID_LEVELS:
            raise EvalParseError(f"row {i}: level must be one of {sorted(VALID_LEVELS)}")
        raw_target = row.get("target_id", "")
        try:
            row_target = str(UUID(raw_target))
        except ValueError:
            raise EvalParseError(f"row {i}: target_id {raw_target!r} is not a uuid") from None
        if level is None:
            level, target_id = row_level, row_target
        elif (row_level, row_target) != (level, target_id):
            raise EvalParseError(
                f"row {i}: a dataset evaluates a single target — "
                f"found {row_level}:{row_target} after {level}:{target_id}"
            )
        if not row.get("input"):
            raise EvalParseError(f"row {i}: input must not be empty")
        grader = row.get("grader") or "llm_judge"
        if grader not in VALID_GRADERS:
            raise EvalParseError(f"row {i}: grader must be one of {sorted(VALID_GRADERS)}")
        cases.append(
            {
                "input": row["input"],
                "expected": row.get("expected", ""),
                "judge_notes": row.get("judge_notes", ""),
                "grader": grader,
            }
        )
    assert level is not None and target_id is not None
    return {"level": level, "target_id": target_id, "cases": cases}
